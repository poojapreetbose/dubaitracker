
"""
ETL (Extract, Transform, Load) for the Dubai real-estate transactions dataset.

Outputs:
- A cleaned dataset (CSV/Parquet/Pickle depending on installed engines)
- A consistent schema for the Streamlit dashboard

Design goals:
- Deterministic, repeatable cleaning
- Safe defaults (never silently delete large chunks without flagging)
- Fast reload after daily updates (raw Excel can be overwritten/extended)
"""
from __future__ import annotations

import os
from dataclasses import asdict
from typing import Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from . import config as C
from .fuzzy import load_alias_map, normalize_developer


def read_excel_fast(xlsx_path: str, sheet_name: str = C.DEFAULT_SHEET_NAME) -> pd.DataFrame:
    """
    Faster-than-pandas read for large XLSX exports by streaming worksheet values.

    If you prefer pandas.read_excel, you can swap this out.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = ws.values
    header = next(rows)
    data = list(rows)
    df = pd.DataFrame(data, columns=list(header))
    return df


def _to_numeric(series: pd.Series) -> pd.Series:
    # Handles numbers that may come in as strings with commas.
    return pd.to_numeric(series.astype(str).str.replace(",", "").str.strip(), errors="coerce")


def map_transaction_bucket(tx: object) -> str:
    if tx is None or (isinstance(tx, float) and np.isnan(tx)):
        return C.TX_BUCKET_OTHER
    s = str(tx).strip().lower()
    # Guard: sometimes exports include a footer like "Applied filters: ..."
    if s.startswith("applied filters"):
        return C.TX_BUCKET_OTHER
    if "off-plan" in s or "off plan" in s:
        return C.TX_BUCKET_OFFPLAN
    if "ready" in s:
        return C.TX_BUCKET_READY
    return C.TX_BUCKET_OTHER


def clean_transactions(
    df_raw: pd.DataFrame,
    developer_alias_csv: Optional[str] = None,
) -> Tuple[pd.DataFrame, dict]:
    """
    Returns (clean_df, audit_dict)
    """
    audit = {}
    df = df_raw.copy()

    # Trim column names
    df.columns = [str(c).strip() for c in df.columns]

    # Drop obvious non-data rows
    before = len(df)
    df = df[df[C.COL_DATE].notna()].copy()
    audit["dropped_rows_missing_date"] = int(before - len(df))

    before = len(df)
    df = df[~df[C.COL_TRANSACTION_TYPE].astype(str).str.lower().str.startswith("applied filters", na=False)].copy()
    audit["dropped_rows_applied_filters_footer"] = int(before - len(df))

    # Standardize strings
    for col in [C.COL_TRANSACTION_TYPE, C.COL_COMMUNITY, C.COL_PROPERTY, C.COL_PROPERTY_TYPE, C.COL_BEDROOMS, C.COL_DEVELOPER]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
            df.loc[df[col].isin(["None", "nan", "NaN", ""]), col] = np.nan

    # Transaction bucket
    df["Transaction Bucket"] = df[C.COL_TRANSACTION_TYPE].apply(map_transaction_bucket)

    # Property type filter (keep only Villa / Apartment)
    before = len(df)
    df = df[df[C.COL_PROPERTY_TYPE].isin(C.ALLOWED_PROPERTY_TYPES)].copy()
    audit["dropped_rows_property_type_not_allowed"] = int(before - len(df))

    # Bedrooms cleaning / filtering
    df[C.COL_BEDROOMS] = df[C.COL_BEDROOMS].astype(str).str.replace(r"\s+", " ", regex=True).str.strip().str.upper()
    df.loc[df[C.COL_BEDROOMS].isin(["NONE", "NAN"]), C.COL_BEDROOMS] = np.nan
    before = len(df)
    df = df[df[C.COL_BEDROOMS].isin(C.ALLOWED_BEDROOMS)].copy()
    audit["dropped_rows_bedrooms_not_allowed"] = int(before - len(df))

    # Times Sold
    if C.COL_TIMES_SOLD in df.columns:
        df[C.COL_TIMES_SOLD] = pd.to_numeric(df[C.COL_TIMES_SOLD], errors="coerce")
        # Many off-plan rows have NaN; interpret as 0 (never sold before / first sale)
        df[C.COL_TIMES_SOLD] = df[C.COL_TIMES_SOLD].fillna(0).clip(lower=0, upper=20).astype(int)

    # Numeric columns
    for col in [C.COL_SIZE_SQF, C.COL_AMOUNT_AED, C.COL_AED_PSF]:
        if col in df.columns:
            df[col] = _to_numeric(df[col])

    # Basic sanity: remove non-positive size/amount
    before = len(df)
    df = df[(df[C.COL_SIZE_SQF] > 0) & (df[C.COL_AMOUNT_AED] > 0)].copy()
    audit["dropped_rows_nonpositive_size_or_amount"] = int(before - len(df))

    # De-duplication guard (useful when new rows are pasted in daily)
    # We intentionally do NOT assume the exported "No" is unique (it is not).
    # Instead, we hash a stable combination of fields that makes a row effectively unique.
    key_cols = [C.COL_DATE, "No", C.COL_PROPERTY, "Unit", C.COL_SIZE_SQF, C.COL_AMOUNT_AED]
    # Some exports may have missing columns (defensive)
    key_cols = [c for c in key_cols if c in df.columns]
    if key_cols:
        df["Transaction Key"] = pd.util.hash_pandas_object(df[key_cols], index=False).astype("uint64").astype(str)
        before = len(df)
        df = df.drop_duplicates(subset=["Transaction Key"]).copy()
        audit["dropped_duplicate_rows_by_key"] = int(before - len(df))


    # Developer normalization (raw + grouped)
    df["Developer Raw"] = df[C.COL_DEVELOPER].fillna("Unknown").astype(str).str.strip()
    alias_map = load_alias_map(developer_alias_csv) if developer_alias_csv else {}

    # IMPORTANT: normalize per-unique value (fast). Do NOT fuzzy-match row-by-row.
    unique_devs = df["Developer Raw"].dropna().unique().tolist()
    dev_map = {
        d: normalize_developer(
            d,
            canonical_names=C.CANONICAL_DEVELOPERS,
            alias_map=alias_map,
            threshold=C.DEVELOPER_FUZZY_THRESHOLD,
        )
        for d in unique_devs
    }
    df["Developer Group"] = df["Developer Raw"].map(dev_map).fillna("Unknown")

    # Community / Property basic standardization
    df[C.COL_COMMUNITY] = df[C.COL_COMMUNITY].fillna("Unknown").astype(str).str.strip()
    df[C.COL_PROPERTY] = df[C.COL_PROPERTY].fillna("Unknown").astype(str).str.strip()

    # Recompute AED/Sqf if missing or clearly incorrect (optional; conservative)
    # Here, we only fill if AED/Sqf is NaN.
    mask_missing_psf = df[C.COL_AED_PSF].isna()
    if mask_missing_psf.any():
        df.loc[mask_missing_psf, C.COL_AED_PSF] = df.loc[mask_missing_psf, C.COL_AMOUNT_AED] / df.loc[mask_missing_psf, C.COL_SIZE_SQF]

    # Date sorting newest -> oldest
    df = df.sort_values(C.COL_DATE, ascending=False).reset_index(drop=True)

    # Helpful derived fields
    df["Date Only"] = pd.to_datetime(df[C.COL_DATE]).dt.date
    df["Month"] = pd.to_datetime(df[C.COL_DATE]).dt.to_period("M").astype(str)
    df["Week"] = pd.to_datetime(df[C.COL_DATE]).dt.to_period("W").astype(str)

    audit["rows_final"] = int(len(df))
    audit["date_min"] = str(pd.to_datetime(df[C.COL_DATE]).min().date())
    audit["date_max"] = str(pd.to_datetime(df[C.COL_DATE]).max().date())
    return df, audit


def add_outlier_flags(df: pd.DataFrame, k_low: float = 1.5, k_high: float = 3.0) -> pd.DataFrame:
    """
    Adds boolean flags marking outliers for Size, Amount, and PSF using a robust IQR rule.

    We do NOT drop rows here; we add flags so the dashboard can exclude outliers by default.
    """
    out = df.copy()

    def flag_iqr(col: str) -> pd.Series:
        x = out[col].dropna()
        if len(x) < 10:
            return pd.Series(False, index=out.index)
        q1 = x.quantile(0.25)
        q3 = x.quantile(0.75)
        iqr = q3 - q1
        lo = q1 - k_low * iqr
        hi = q3 + k_high * iqr
        return (out[col] < lo) | (out[col] > hi)

    out["Outlier Size"] = flag_iqr(C.COL_SIZE_SQF)
    out["Outlier Amount"] = flag_iqr(C.COL_AMOUNT_AED)
    out["Outlier PSF"] = flag_iqr(C.COL_AED_PSF)
    out["Any Outlier"] = out[["Outlier Size", "Outlier Amount", "Outlier PSF"]].any(axis=1)

    return out


def save_processed(df: pd.DataFrame, out_path: str) -> None:
    """
    Saves processed dataset in a sensible format based on extension.
    """
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    ext = os.path.splitext(out_path)[1].lower()

    if ext == ".parquet":
        # Requires pyarrow or fastparquet
        df.to_parquet(out_path, index=False)
    elif ext == ".pkl":
        df.to_pickle(out_path)
    elif ext in [".csv", ".gz"]:
        df.to_csv(out_path, index=False)
    else:
        # default fallback
        df.to_pickle(out_path)


def load_processed(in_path: str) -> pd.DataFrame:
    ext = os.path.splitext(in_path)[1].lower()
    if ext == ".parquet":
        return pd.read_parquet(in_path)
    if ext == ".pkl":
        return pd.read_pickle(in_path)
    return pd.read_csv(in_path)


def etl_run(
    raw_xlsx_path: str,
    out_path: str,
    developer_alias_csv: Optional[str] = None,
    outlier_cfg: Optional[C.OutlierConfig] = None,
) -> dict:
    """
    End-to-end run. Returns audit dict (log-friendly).

    Typical usage:
        audit = etl_run("data/raw/transactions.xlsx", "data/processed/transactions_clean.pkl")
    """
    df_raw = read_excel_fast(raw_xlsx_path, sheet_name=C.DEFAULT_SHEET_NAME)
    df_clean, audit = clean_transactions(df_raw, developer_alias_csv=developer_alias_csv)

    outlier_cfg = outlier_cfg or C.DEFAULT_OUTLIER_CONFIG
    if outlier_cfg.enabled:
        df_clean = add_outlier_flags(df_clean, k_low=outlier_cfg.k_low, k_high=outlier_cfg.k_high)

    save_processed(df_clean, out_path)
    audit["output_path"] = out_path
    audit["outlier_config"] = asdict(outlier_cfg)
    return audit
